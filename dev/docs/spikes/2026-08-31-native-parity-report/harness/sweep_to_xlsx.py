#!/usr/bin/env python3
"""Adapter: loads a `sweep_corpus.py` JSON result file (`sweep_lib.SweepRun`) and writes/replaces one
sheet in the shared tracking workbook via `openpyxl`. Deliberately separate from `sweep_corpus.py` --
the sweep never imports `openpyxl` and this script never spawns an editor, so a future change to
either (a new xlsx column, a new sweep field) doesn't require touching both.

Column layout matches the existing `2026-09-02_0957Z` sheet exactly (checked live before writing this):
level, dx_path, status, nodes_match, surfs_match, leaves_match, verts_match, points_match,
vectors_match, geometry_match_count, content_exact_fraction, lighting_byte_identical_pct,
lighting_shadow_bit_pct, full_parity, notes -- 15 columns, A-O. `content_exact_fraction` is the
index-for-index STRUCTURAL content match (not just the six geometry array-length counts); a level can
score geometry_match_count=6 while this is well below 1.0, which is the exact case a
length/count-only report would hide (see `sweep_lib.content_exact_fraction`'s docstring).

Usage:
  .venv/bin/python sweep_to_xlsx.py <sweep.json> --workbook FILE --sheet-name NAME
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

import sweep_lib as sl  # noqa: E402

COLUMNS = (
    "level", "dx_path", "status", "nodes_match", "surfs_match", "leaves_match", "verts_match",
    "points_match", "vectors_match", "geometry_match_count", "content_exact_fraction",
    "lighting_byte_identical_pct", "lighting_shadow_bit_pct", "full_parity", "notes",
)


def level_result_row(r: sl.LevelResult) -> tuple:
    return tuple(getattr(r, col) for col in COLUMNS)


def write_sheet(run: sl.SweepRun, workbook_path: Path, sheet_name: str) -> None:
    import openpyxl

    if workbook_path.exists():
        wb = openpyxl.load_workbook(workbook_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # drop the default blank sheet -- every real sheet is named explicitly

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(COLUMNS)
    for r in run.results:
        ws.append(level_result_row(r))

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0],
                                 formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__)
    ap.add_argument("sweep_json", help="path to a sweep_corpus.py JSON output file")
    ap.add_argument("--workbook", required=True, help="path to the .xlsx workbook to write/update")
    ap.add_argument("--sheet-name", default=None,
                    help="sheet name (default: the sweep's own started_at timestamp, "
                         "YYYY-MM-DD_HHMMZ -- matching the existing '2026-09-02_0957Z' sheet naming)")
    args = ap.parse_args(argv)

    sweep_path = Path(args.sweep_json)
    if not sweep_path.is_file():
        print(f"sweep_to_xlsx: not a file: {sweep_path}", file=sys.stderr)
        return 2
    run = sl.read_sweep_json(sweep_path)

    # started_at is "YYYY-MM-DDTHH:MM:SS+00:00" -- [:16] is "YYYY-MM-DDTHH:MM"; matches the existing
    # sheet-naming convention "YYYY-MM-DD_HHMMZ".
    sheet_name = args.sheet_name or run.started_at[:16].replace("T", "_").replace(":", "") + "Z"
    write_sheet(run, Path(args.workbook), sheet_name)
    print(f"[sweep_to_xlsx] wrote sheet {sheet_name!r} ({len(run.results)} rows) to {args.workbook}",
         file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
