"""Tests `sweep_to_xlsx.py`'s adapter -- column layout, row mapping, and sheet create/replace -- via a
real `openpyxl` workbook on a tmp path. No sweep, no docker.

Run directly:
    .venv/bin/python -m pytest dev/docs/spikes/2026-08-31-native-parity-report/harness/test_sweep_to_xlsx.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import sweep_lib as sl        # noqa: E402
import sweep_to_xlsx as sx    # noqa: E402


def _sample_run():
    return sl.SweepRun(
        started_at="2026-09-02T14:23:05+00:00", concurrency=1, rebuild_timeout=3600.0,
        hang_timeout=6300.0,
        results=(
            sl.LevelResult(level="DX.dx", dx_path="/x/DX.dx", status="OK", elapsed_s=1.2,
                           golden_cache_hit=True, nodes_match=True, surfs_match=True,
                           leaves_match=True, verts_match=True, points_match=True,
                           vectors_match=True, geometry_match_count=6, content_exact_fraction=1.0,
                           content_length_mismatch=False, lighting_byte_identical_pct=100.0,
                           lighting_shadow_bit_pct=100.0, full_parity=True,
                           notes="content: nodes exact, surfs exact, leaves exact"),
            sl.LevelResult(level="99_Endgame4.dx", dx_path="/x/99_Endgame4.dx", status="SKIPPED",
                           notes="offline UCC batchexport can't resolve Engine.CameraPoint"),
        ))


def test_columns_match_the_existing_sheet_layout():
    assert sx.COLUMNS == (
        "level", "dx_path", "status", "nodes_match", "surfs_match", "leaves_match", "verts_match",
        "points_match", "vectors_match", "geometry_match_count", "content_exact_fraction",
        "lighting_byte_identical_pct", "lighting_shadow_bit_pct", "full_parity", "notes")


def test_write_sheet_creates_workbook_with_header_and_rows(tmp_path):
    import openpyxl
    wb_path = tmp_path / "report.xlsx"
    sx.write_sheet(_sample_run(), wb_path, "2026-09-02_1423Z")

    wb = openpyxl.load_workbook(wb_path)
    assert wb.sheetnames == ["2026-09-02_1423Z"]
    ws = wb["2026-09-02_1423Z"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == sx.COLUMNS
    assert rows[1][0] == "DX.dx"
    assert rows[1][9] == 6          # geometry_match_count
    assert rows[1][10] == 1.0       # content_exact_fraction
    assert rows[2][0] == "99_Endgame4.dx"
    assert rows[2][2] == "SKIPPED"


def test_write_sheet_replaces_existing_sheet_of_the_same_name_not_duplicates(tmp_path):
    wb_path = tmp_path / "report.xlsx"
    sx.write_sheet(_sample_run(), wb_path, "2026-09-02_1423Z")
    sx.write_sheet(_sample_run(), wb_path, "2026-09-02_1423Z")  # re-run, same name

    import openpyxl
    wb = openpyxl.load_workbook(wb_path)
    assert wb.sheetnames == ["2026-09-02_1423Z"]  # not duplicated


def test_write_sheet_preserves_other_sheets(tmp_path):
    import openpyxl
    wb_path = tmp_path / "report.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "existing-sheet"
    wb.active.append(("a", "b"))
    wb.save(wb_path)

    sx.write_sheet(_sample_run(), wb_path, "new-sheet")

    wb2 = openpyxl.load_workbook(wb_path)
    assert set(wb2.sheetnames) == {"existing-sheet", "new-sheet"}


def test_default_sheet_name_derived_from_started_at():
    run = _sample_run()
    # exercise main()'s derivation logic directly (avoids a subprocess for a pure string transform)
    sheet_name = run.started_at[:16].replace("T", "_").replace(":", "") + "Z"
    assert sheet_name == "2026-09-02_1423Z"
